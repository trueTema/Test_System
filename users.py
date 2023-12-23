import datetime
import json
import os
import sqlite3
import time
from telebot import types
import run_checker
import LinkedList
import database
import telebot as tgbot
from Task import TASK
from Parcel import Parcel
import threading
from errors import errors
from openpyxl import Workbook

_token = json.load(open('Files/config.json', 'r'))["token"]
bot = tgbot.TeleBot(_token)

statuses = {
    "student": 0,
    "teacher": 1,
    "super_user": 2
}

inf = 4133969999

key_word = "8pJSSMgH7B"  # KeyWord for admin


current_parcel_number = 1


banned_users = set()
"""Blocked users"""

"""
Admin login will be initialized with a hiden command - /adminLog key_word
"""

_help = "Список доступных команд:\n\n" \
        "help - просмотр списка команд\n" \
        "send <ID> - отправить решение на проверку\n" \
        "status <ID> - получить информацию о своих текущих баллах за задачу\n" \
        "profile - получить информацию о своём профиле\n" \
        "getTask <ID> - получить задачу по ID\n" \
        "\n===============\nКоманды для учителя:\n" \
        "updateTask <ID> - обновить задачу\n" \
        "deleteTask <ID> - удалить задачу\n" \
        "addTask <id> <open\hide> <best\last> [dd.MM.YYYY-HH:mm (Время дедлайна)] [Группа] - добавить задачу\n" \
        "exit - смена статуса на ученика\n" \
        "stats <ID> - получить статистику по задаче\n" \
        "getLog <ID> - получить полный лог посылок по задаче\n" \
        "\n===============\nКоманды для суперпользователя:\n" \
        "ban <ID> - заблокировать пользователя\n" \
        "unban <ID> - разблокировать пользователя\n" \
        "changeName/changeStatus/changeGroup <ID> <Имя/Статус/Группа> - изменить имя, статус, группу пользователя\n"
"""
List of commands for teacher which will be in his own "special" help

Some description for Super User commands
Commands like 'add/deleteTask' don't need to be explained
'add\deleteGroup' needs for cases like: User enters group, which doesn't exist and we throw error\exception
'getTask' is for correct the fields of chosen task ( SuperUser find out that, for example, field answer is wrong and has
                                                                                        to be corrected
"""

connections = dict()
"""Hash table that contains connections for threads"""


def get_connection(thread_id: int) -> sqlite3.Connection:
    """
    returns connection to database for any thread
    :param thread_id:
    :return: connection to database
    """
    if thread_id in connections.keys():
        return connections[thread_id]
    connection = sqlite3.connect("Files/database.sql")
    connections[thread_id] = connection
    return connection


class User:
    _cmd_status = None

    """
    Needed in order to enter the wording of the task itself into the Task class as a separate message
    After adding the wording to the class, it is "pushed" to the database and again becomes None
    """

    cur_Task = None
    cur_task_id = None

    def __init__(self, id: int, name='', status: str = "student",
                 study_group=None):
        """Initializing constructor"""
        self.id = id
        self.name = name
        self.status = status
        self.study_group = study_group

    def doc_handler(self, file: bytes, filetype: str):
        if self._cmd_status == 'pushing_checking_script':
            if filetype != 'py':
                bot.send_message(self.id, 'Скрипт должен быть формата .py')
                return
            src = f"Scripts/{str(self.cur_Task.id)}.py"
            with open(src, "wb") as new_file:
                new_file.write(file)
                new_file.close()
            connection = get_connection(threading.current_thread().native_id)
            # bot.send_message(self.id,self.cur_Task.id)
            # bot.send_message(self.id, self.cur_Task.statement)
            # bot.send_message(self.id, str(self.cur_Task.deadline))
            database.add_problem(self.cur_Task, connection)
            self._cmd_status = None
            self.cur_Task = None
            bot.send_message(self.id, 'Задача успешно добавлена.')
            return
        if self._cmd_status == 'updating_checking_script':
            if filetype != 'py':
                bot.send_message(self.id, 'Скрипт должен быть формата .py')
                self.cur_Task = None
                self._cmd_status = ''
                return
            os.remove(f'Scripts/{self.cur_Task.id}.py')
            src = f"Scripts/{str(self.cur_Task.id)}.py"
            with open(src, "wb") as new_file:
                new_file.write(file)
                new_file.close()
            connection = get_connection(threading.current_thread().native_id)
            database.update_problem(self.cur_Task, connection)
            bot.send_message(self.id, "Задача была успешно обновлена.")
            self.cur_Task = None
            self._cmd_status = None
            return
        if self._cmd_status == "pushing_parcel":
            if filetype != 'txt':
                bot.send_message(self.id, 'Посылкой может быть только текстовый файл.')
                return
            now = datetime.datetime.now()
            timestamp = int(datetime.datetime.timestamp(now))
            try:
                checker_result = run_checker.get_points(f'Scripts/{self.cur_task_id}.py',
                                                        run_checker.bytes_to_str(file))
            except Exception as e:
                bot.send_message(self.id, 'Нетестируемый ответ.')
                return
            if len(checker_result["errors"]) != 0:
                bot.send_message(self.id, checker_result["errors"])
                return
            if database.get_problem(self.cur_task_id,
                                    get_connection(threading.current_thread().native_id)).visible == 1:
                bot.send_message(self.id, f'Набрано баллов: {checker_result["points"]}')

            current_parcell = Parcel(points=checker_result["points"],
                                     date=timestamp, id_user=self.id, id_task=self.cur_task_id,
                                     answer=run_checker.bytes_to_str(file), id=int(str(self.id)[1::3] + str(now)[-6:]))
            connection = get_connection(threading.current_thread().native_id)
            database.add_parcel(current_parcell, connection)  # Sending parcell
            self.cur_Task = None
            self._cmd_status = None
            bot.send_message(self.id, "Посылка была отправлена")
            return

    def txt_handler(self, txt: str):
        """
        Text messages handler for each user.
        :param txt: message text
        """
        if self.id in banned_users:
            bot.send_message(self.id, 'Вы заблокированы.')
            return
        if self._cmd_status == "register_r_name":
            while True:
                if len(txt.split(" ")) != 2:
                    bot.send_message(self.id, 'Некорректный формат Фамилии Имя')
                    return
                else:
                    self.name = txt
                    break
            bot.send_message(self.id, 'Введите учебную группу: ')
            self._cmd_status = "register_r_study_g"
            return
        if self._cmd_status == "pushing_parcel":
            now = datetime.datetime.now()
            timestamp = int(datetime.datetime.timestamp(now))
            try:
                checker_result = run_checker.get_points(f'Scripts/{self.cur_task_id}.py',
                                                        txt)
            except Exception as e:
                bot.send_message(self.id, 'Нетестируемый ответ.')
                self._cmd_status = ""
                return
            if len(checker_result["errors"]) != 0:
                bot.send_message(self.id, checker_result["errors"])
                self._cmd_status = ""
                return
            if database.get_problem(self.cur_task_id,
                                    get_connection(threading.current_thread().native_id)).visible == 1:
                bot.send_message(self.id, f'Набрано баллов: {checker_result["points"]}')

            current_parcell = Parcel(points=checker_result["points"],
                                     date=timestamp, id_user=self.id, id_task=self.cur_task_id,
                                     answer=txt, id=int(str(self.id)[1::3] + str(now)[-6:]))
            connection = get_connection(threading.current_thread().native_id)
            database.add_parcel(current_parcell, connection)  # Sending parcell
            self.cur_Task = None
            self._cmd_status = None
            bot.send_message(self.id, "Посылка была отправлена")
            return
        if self._cmd_status == "register_r_study_g":
            self.study_group = txt
            bot.send_message(self.id, f'Регистрация успешно завершена. Добро пожаловать, {self.name}!')
            self._cmd_status = None
            return

        if self._cmd_status == "admin_pulling_task":
            field_text = txt
            self.cur_Task.statement = field_text
            """
            Adding the number sending time is necessary 
            for the possible implementation of the deadline system in the future
            """
            self._cmd_status = 'pushing_checking_script'
            bot.send_message(self.id, 'Отправьте скрипт проверки для этой задачи.')
            return
        if self._cmd_status == "updating_checker_status":
            field_text = txt
            if field_text == "No":
                self._cmd_status = "updating_of_deadline"
                bot.send_message(self.id,
                                 "Вы хотите обновить дату дедлайна? Если нет - напишите No, в противном случае - новую дату")
                return
            elif field_text == "last":
                new_status = 1
            elif field_text == 'best':
                new_status = 0
            else:
                bot.send_message(self.id, 'Некорректный формат параметра видимости результата посылки.')
                self._cmd_status = ""
                return
            self.cur_Task.best_or_last = new_status
            self._cmd_status = "updating_of_deadline"
            bot.send_message(self.id,
                             "Вы хотите обновить дату дедлайна? Если нет - напишите No, в противном случае - новую дату")
            return
        if self._cmd_status == "updating_of_visibility":
            field_text = txt
            if field_text == "No":
                self._cmd_status = "updating_checker_status"
                bot.send_message(self.id,
                                 "Вы хотите обновить статус чекера? Если нет - напишите No, в противном случае - новый статус")
                return
            elif field_text == "open":
                new_visible = 1
            elif field_text == 'hide':
                new_visible = 0
            else:
                bot.send_message(self.id, 'Некорректный формат параметра видимости результата посылки.')
                self._cmd_status = ""
                return
            self.cur_Task.visible = new_visible
            self._cmd_status = "updating_of_deadline"
            bot.send_message(self.id,
                             "Вы хотите обновить статус чекера? Если нет - напишите No, в противном случае - новый статус")
            return

        if self._cmd_status == "updating_of_deadline":
            field_text = txt
            if field_text == "No":
                bot.send_message(self.id,
                                 "Вы хотите обновить\добавить группу? Если нет - напишите No, в противном случае - новую группу")
                self._cmd_status="updating_of_group"
                return
            else:
                if field_text == '_':
                    date_of_death = datetime.datetime.fromtimestamp(inf)
                else:
                    date_of_death = datetime.datetime.strptime(field_text, "%d.%m.%Y-%H:%M")
                now = datetime.datetime.now()
                if now > date_of_death:
                    bot.send_message(self.id, "Вы пытаетесь установить невозможный дедлайн")
                    self._cmd_status = ""
                    return
                else:
                    timestamp_sec = int(datetime.datetime.timestamp(date_of_death))
                    self.cur_Task.deadline = timestamp_sec
                    bot.send_message(self.id,
                                     "Вы хотите обновить\добавить группу? Если нет - напишите No, в противном случае - новую группу")
                    self._cmd_status = "updating_of_group"
                    return

        if self._cmd_status == "updating_of_group":
            field_text = txt
            bot.send_message(self.id,
                             "Вы хотите обновить условие задачи? Если нет - напишите No, в противном случае - новое условие")
            self._cmd_status = "updating_of_statement"
            if field_text == 'No':
                return
            self.cur_Task.group=field_text
            return

        if self._cmd_status == "updating_of_statement":
            field_text = txt
            bot.send_message(self.id, "Вы хотите обновить скрипт проверки? Если нет - "
                                      "напишите No, в противном случае - Yes. ")
            self._cmd_status = "updating_checking_script_answer"
            if field_text == 'No':
                return
            self.cur_Task.statement = field_text
            return
        if self._cmd_status == "updating_checking_script_answer":
            field_text = txt
            if field_text == 'No':
                connection = get_connection(threading.current_thread().native_id)
                database.update_problem(self.cur_Task, connection)
                bot.send_message(self.id, "Задача была успешно обновлена")
                self.cur_Task = None
                self._cmd_status = None
                return
            elif field_text == 'Yes':
                bot.send_message(self.id, 'Отправьте новый скрипт.')
                self._cmd_status = "updating_checking_script"
                return
            else:
                self.cur_Task = None
                self._cmd_status = None
                return

    def super_user_cmd(self, cmd: str):
        """
        Super user commands handler
        :param cmd: command
        """

        cmd = cmd.split()
        if cmd[0] == 'changeStatus':
            user_id = int(cmd[1])
            new_status = cmd[2]
            if new_status not in statuses.keys():
                bot.send_message(self.id, 'Неизвестный статус пользователя.')
                return
            if user_id in cache.keys():
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять статус суперпользователя.')
                    return
                cache[user_id].data[1].status = new_status
                bot.send_message(self.id, f'Статус пользователя {user_id} успешно изменён на {new_status}')
            else:
                update_cache(user_id, get_connection(threading.current_thread().native_id))
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять статус суперпользователя.')
                    return
                cache[user_id].data[1].status = new_status
                bot.send_message(self.id, f'Статус пользователя {user_id} успешно изменён на {new_status}')
            return
        if cmd[0] == 'changeName':
            user_id = int(cmd[1])
            new_name = ' '.join(cmd[2:])
            if user_id in cache.keys():
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять имя суперпользователя.')
                    return
                cache[user_id].data[1].name = new_name
                bot.send_message(self.id, f'Имя пользователя {user_id} успешно изменёно на {new_name}')
            else:
                update_cache(user_id, get_connection(threading.current_thread().native_id))
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять имя суперпользователя.')
                    return
                cache[user_id].data[1].name = new_name
            return
        if cmd[0] == 'changeGroup':
            user_id = int(cmd[1])
            new_group = cmd[2]
            if user_id in cache.keys():
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять группу суперпользователя.')
                    return
                cache[user_id].data[1].study_group = new_group
                bot.send_message(self.id, f'Группа пользователя {user_id} успешно изменёна на {new_group}')
            else:
                update_cache(user_id, get_connection(threading.current_thread().native_id))
                if cache[user_id].data[1].status == 'super_user' and (self.id != 451938981) and (user_id != self.id):
                    bot.send_message(self.id, 'Невозможно поменять группу суперпользователя.')
                    return
                cache[user_id].data[1].study_group = new_group
            return
        if cmd[0] == 'getUserList':
            if len(cmd) != 1:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return

            user_list = database.get_user_list(get_connection(threading.current_thread().native_id))
            wb = Workbook()

            active_sheet = wb.active
            active_sheet.append(["ID пользователя", "ФИ", "Статус", "Учебная группа"])
            for line in user_list:
                active_sheet.append(
                    [line.id, line.name, line.status, line.study_group])
            wb.save(f'Files/USER_LIST.xlsx')
            wb.close()
            with open(f"Files/USER_LIST.xlsx", "rb") as misc:
                bot.send_message(self.id, 'Таблица успешно создана.')
                bot.send_document(self.id, misc)
            os.remove(f'Files/USER_LIST.xlsx')
            return
        if cmd[0] == 'ban':
            if len(cmd) > 2:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return
            try:
                user_id = int(cmd[1])
            except Exception as e:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return
            if user_id in cache.keys():
                user = cache[user_id].data[1]
            else:
                user = database.get_user(user_id, get_connection(threading.current_thread().native_id))
            if user is not None and user.status == 'super_user' and (self.id != 451938981):
                bot.send_message(self.id, 'Невозможно заблокировать суперпользователя.')
                return
            if user_id == self.id:
                bot.send_message(self.id, 'Невозможно заблокировать себя.')
                return
            banned_users.add(user_id)
            bot.send_message(self.id, 'Пользователь успешно заблокирован.')
            return
        if cmd[0] == 'unban':
            if len(cmd) > 2:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return
            try:
                user_id = int(cmd[1])
            except ...:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return
            if user_id not in banned_users:
                bot.send_message(self.id, 'Пользователь не заблокирован.')
                return
            banned_users.remove(user_id)
            bot.send_message(self.id, 'Пользователь успешно разблокирован.')
            return

    def cmd_handler(self, cmd: str):
        """
        Command handler for each user
        :param cmd: command
        """

        if self.id in banned_users:
            bot.send_message(self.id, 'Вы заблокированы.')
            return
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add(types.KeyboardButton(text="/help"), types.KeyboardButton(text="/status"))
        if cmd[:3] == 'su ':
            if self.status != 'super_user':
                bot.send_message(self.id, "Ошибка доступа.", reply_markup=kb)
                return
            self.super_user_cmd(cmd[3:])
            return
        if cmd[:6] == 'getLog':
            cmd = cmd.split()[1:]
            if len(cmd) != 1:
                bot.send_message(self.id, errors["IncorrectFormat"])
                return
            task_id = int(cmd[0])
            connection = get_connection(threading.current_thread().native_id)
            cur_task = database.get_problem(task_id, connection)
            if cur_task is None:
                bot.send_message(self.id, errors["IDNotFound"])
                return
            if (self.status == 'teacher' and self.id != cur_task.id_of_user) or (self.status == 'student'):
                bot.send_message(self.id, errors["AccessError"])
                return
            parcels_list = database.get_problem_parcels(task_id, connection)
            if len(parcels_list) == 0:
                bot.send_message(self.id, 'По этой задаче ещё не было посылок.')
                return
            wb = Workbook()

            active_sheet = wb.active
            active_sheet.append(["Время отправки", "ID посылки", "ID отправителя", "ФИ отправителя", "Ответ", "Баллы"])
            for line in parcels_list:
                user = database.get_user(line.id_user, connection)
                active_sheet.append([str(datetime.datetime.utcfromtimestamp(line.date + 3*60*60).strftime('%Y-%m-%d %H:%M:%S')),
                           line.id, line.id_user, str(user.name), str(line.answer).encode(), line.points])
            wb.save(f'Files/LOG_{task_id}.xlsx')
            wb.close()
            with open(f"Files/LOG_{task_id}.xlsx", "rb") as misc:
                bot.send_message(self.id, 'Таблица успешно создана.')
                bot.send_document(self.id, misc)
            os.remove(f'Files/LOG_{task_id}.xlsx')
            return
        if cmd == 'start':
            if self.name != '':
                bot.send_message(self.id, "Вы уже зарегистрированы.", reply_markup=kb)
                return
            bot.send_message(self.id, 'Введите Фамилию Имя:')
            self._cmd_status = "register_r_name"
            return
        if self.name == '':
            bot.send_message(self.id, 'Вы не завершили регистрацию.')
            return
        if cmd == 'help':
            bot.send_message(self.id, _help)
            return
        if cmd[:4] == 'send':
            if len(cmd.split(" ")) != 2:
                bot.send_message(self.id, "Вы не ввели айди задачи или сделали это неправильно ")
                return
            id_of_parcell = int(cmd.split(" ")[1])
            connection = get_connection(threading.current_thread().native_id)
            if database.get_problem(id_of_parcell, connection) is None:
                bot.send_message(self.id, "Вы пытаетесь отправить решение для несуществующей задачи")
                return
            now = datetime.datetime.now()
            chosen_task = database.get_problem(id_of_parcell, connection)
            time_of_task = datetime.datetime.fromtimestamp(chosen_task.deadline)
            if time_of_task < now:
                bot.send_message(self.id, "Ответы для данной задачи не принимаются")
                return
            self._cmd_status = "pushing_parcel"
            self.cur_task_id = id_of_parcell
            bot.send_message(self.id, "Отправьте файл с ответом.", reply_markup=kb)
            return
        if cmd[:8] == "adminLog":  # Login as admin
            if statuses[self.status] == 2:
                bot.send_message(self.id,"Вы уже имеете статус учителя")
                return
            if str(cmd[9:]) != "" and str(cmd[9:]) == key_word:  # !!! The key_word has to be right !!!
                kb1 = kb.add(types.KeyboardButton(text="/adminHelp"))
                bot.send_message(self.id, f"Добро пожаловать, {self.name}!",
                                 reply_markup=kb1)  # Special keyboard for admin
                self.status = "super_user"
            else:
                bot.send_message(self.id, str(cmd[9:]))  # Exception
                bot.send_message(self.id, "Ошибка доступа.")
            return
        if cmd[:10] == "updateTask":
            if statuses[self.status] == 0:
                bot.send_message(self.id, "Ошибка доступа.")
                return
            else:
                if len(cmd.split()) > 2:
                    bot.send_message(self.id, "Команда должна быть в виде /updateTask <ID>")
                    return
                if not cmd.split(" ")[1].isdigit():
                    bot.send_message(self.id, "ID должно быть числом")
                    return
                id_of_task = int(cmd.split(" ")[1])
                connection = get_connection(threading.current_thread().native_id)
                problem = database.get_problem(id_of_task, connection)
                if problem is None:
                    bot.send_message(self.id, "Вы пытаетесь обновить несуществующую задачу")
                    return
                current_task = database.get_problem(id_of_task, connection)
                self.cur_Task = current_task
                self._cmd_status = "updating_of_visibility"
                bot.send_message(self.id, "Вы хотите обновить видимость? Если нет - напишите No, в противном случае - новую видимость")
                return
        if cmd == 'profile':
            bot.send_message(self.id, f'Имя: {self.name}\n'
                                      f'Статус: {self.status}\n'
                                      f'Группа: {self.study_group}'
                             )
            return
        if cmd == "exit":
            if self.status == "teacher" or self.status == "super_user":
                bot.send_message(self.id, "Ваш статус был изменен", reply_markup=kb)  # super user to student)
                self.status = "student"
                return
            else:
                bot.send_message(self.id, "У вас нет доступа к данной команде")
        if cmd[:10] == "deleteTask":
            if self.status == "teacher" or self.status == "super_user":
                if len(cmd.split(" ")) == 1:
                    bot.send_message(self.id, "Вы не ввели ID задачи")
                    return
                if not cmd.split(" ")[1].isdigit():
                    bot.send_message(self.id, "ID должно быть числом")
                    return
                id_of_task = int(cmd.split(" ")[1])
                connection = get_connection(threading.current_thread().native_id)
                problem = database.get_problem(id_of_task, connection)
                if problem is not None:
                    if problem.id_of_user != self.id and self.status == 'teacher':
                        bot.send_message(self.id, "Ошибка доступа.")
                        return
                    database.delete_problem(id_of_task, connection)
                    os.remove(f'Scripts/{id_of_task}.py')
                    bot.send_message(self.id, f"Задача под индексом {id_of_task} была удалена")
                    return
                else:
                    bot.send_message(self.id, "Такой задачи не найдено.")
                    return
            else:
                bot.send_message(self.id, "Ошибка доступа.")
                return
        if cmd[:6] == 'status':
            cmd = cmd.split()
            if len(cmd) != 2:
                bot.send_message(self.id, 'Некорректный формат.')
                return
            try:
                problem_id = int(cmd[1])
            except Exception as e:
                bot.send_message(self.id, 'Некорректный формат.')
                return
            parcels_list = database.get_user_problem_parcels(self.id, problem_id,
                                                             get_connection(threading.current_thread().native_id))
            if len(parcels_list) == 0:
                bot.send_message(self.id, "У вас ещё нет посылок.")
                return
            is_visible = True if database.get_problem(problem_id, get_connection(
                threading.current_thread().native_id)).visible == 1 else False
            parcels_list.sort(key=lambda x: x.date)
            res = 'Ваши посылки:\n\nНомер\Баллы\Время отправки\n'
            for i in range(len(parcels_list)):
                date = datetime.datetime.utcfromtimestamp(parcels_list[i].date).strftime('%Y-%m-%d %H:%M:%S')
                res += f'[{i + 1}]' + ' ' + f'{parcels_list[i].points if is_visible else "Баллы скрыты"}' + ' ' + f'{date}\n'
            bot.send_message(self.id, res)
        if cmd[:12] == 'deleteParcel':
            if not(self.status == 'teacher' or self.status == 'super_user'):
                bot.send_message(self.id, 'Ошибка доступа.')
                return
            if len(cmd.split()) != 2:
                bot.send_message(self.id, 'Некорректный формат команды')
                return
            parcel_id = int(cmd.split()[-1])
            connection = get_connection(threading.current_thread().native_id)
            parcel = database.get_parcel_by_id(parcel_id, connection)
            if parcel is None:
                bot.send_message(self.id, 'Посылка с таким ID не найдена.')
                return
            teacher_id = database.get_problem(parcel.id_task, connection).id_of_user
            if self.id != teacher_id and self.status == "teacher":
                bot.send_message(self.id, 'Ошибка доступа.')
                return
            database.delete_parcel(parcel_id, connection)
            bot.send_message(self.id, 'Посылка успешно удалена')
        if cmd[:7] == "addTask":
            if self.status == 'student':
                bot.send_message(self.id, 'Ошибка доступа.')
                return
            cmd = cmd.split()[1:]
            try:
                id = int(cmd[0])
            except Exception as e:
                bot.send_message(self.id, 'Некорректный ID задачи.')
                return
            if len(cmd) < 2:
                bot.send_message(self.id, 'Отсутствует параметр видимости результата посылки.')
                return
            if cmd[1] == 'open':
                visibility_status = 1
            elif cmd[1] == 'hide':
                visibility_status = 0
            else:
                bot.send_message(self.id, 'Некорректный параметр видимости результата посылки.')
                return
            if len(cmd) < 3:
                bot.send_message(self.id, 'Отсутствует параметр статуса чекера.')
                return
            if cmd[2] == 'best':
                checker_status = 0
            elif cmd[2] == 'last':
                checker_status = 1
            else:
                bot.send_message(self.id, 'Некорректный параметр статуса чекера.')
                return
            deadline = inf
            if len(cmd) > 3:
                string_of_date = str(cmd[3])
                if string_of_date != '_':
                    try:
                        date_of_death = datetime.datetime.strptime(string_of_date, "%d.%m.%Y-%H:%M")
                    except Exception as e:
                        bot.send_message(self.id, "Некорректный формат дедлайна.")
                        return
                    now = datetime.datetime.now()
                    if now > date_of_death:
                        bot.send_message(self.id, "Вы пытаетесь установить невозможный дедлайн")
                        return
                    deadline = int(datetime.datetime.timestamp(date_of_death))

            group = "None"
            if len(cmd) > 4 and cmd[4] != '_':
                group = cmd[4]
            connection = get_connection(threading.current_thread().native_id)
            cur_task = database.get_problem(id, connection)
            if cur_task is not None:
                bot.send_message(self.id, 'Данная задача уже существует.')
                return
            bot.send_message(self.id, "Отправьте условие задачи.")
            self._cmd_status = "admin_pulling_task"
            new_one = TASK(id=id, visible=visibility_status, group=group, id_of_user=self.id, deadline=deadline,
                           best_or_last=checker_status)
            self.cur_Task = new_one
            return
        if cmd[:5] == "stats":
            if statuses[self.status] == 0:
                bot.send_message(self.id, "Ошибка доступа.")
                return
            connection = get_connection(threading.current_thread().native_id)
            try:
                id_of_task = int(cmd.split(" ")[1])
            except Exception as e:
                bot.send_message(self.id, 'Некорректный формат команды.')
                return
            problem = database.get_problem(id_of_task, connection)
            if problem is None:
                bot.send_message(self.id, 'Задача не найдена.')
                return
            if self.status == 'teacher' and problem.id_of_user != self.id:
                bot.send_message(self.id, "Ошибка доступа.")
            parcels_list = database.get_problem_parcels(id_of_task, connection)
            if len(parcels_list) == 0:
                bot.send_message(self.id, "Для данной задачи не было отправлено ни одного решения")
                return
            if problem.best_or_last == 0:
                parcels_list.sort(key=lambda x: [x.id_user, inf - x.points])
            else:
                parcels_list.sort(key=lambda x: [x.id_user, inf - x.date])

            list_of_chosen_id = [parcels_list[0]]

            for i in range(1, len(parcels_list)):
                if parcels_list[i].id_user != parcels_list[i - 1].id_user:
                    list_of_chosen_id += [parcels_list[i]]
            list_of_chosen_id.sort(key=lambda parcel: [parcel.points, inf - parcel.date], reverse=True)
            wb = Workbook()

            active_sheet = wb.active
            active_sheet.append(["Время отправки", "ID посылки", "ID отправителя", "ФИ отправителя", "Ответ", "Баллы"])
            for line in list_of_chosen_id:
                user = database.get_user(line.id_user, connection)
                active_sheet.append(
                    [str(datetime.datetime.utcfromtimestamp(line.date + 3 * 60 * 60).strftime('%Y-%m-%d %H:%M:%S')),
                     line.id, line.id_user, str(user.name), str(line.answer).encode(), line.points])
            wb.save(f'Files/STATS_{id_of_task}.xlsx')
            wb.close()
            with open(f"Files/STATS_{id_of_task}.xlsx", "rb") as misc:
                bot.send_message(self.id, 'Таблица успешно создана.')
                bot.send_document(self.id, misc)

            os.remove(f'Files/STATS_{id_of_task}.xlsx')
            return
        if cmd[:7] == 'getTask':
            cmd = cmd.split()[1:]
            try:
                task_id = int(cmd[0])
            except Exception as e:
                bot.send_message(self.id, 'Некорректный формат ввода.')
                return
            cur_task = database.get_problem(task_id, get_connection(threading.current_thread().native_id))
            if cur_task is None:
                bot.send_message(self.id, 'Задача не существует.')
                return
            if (cur_task.id_of_user != self.id and cur_task.group != self.study_group and self.status != 'super_user') and cur_task.group != "None":
                bot.send_message(self.id, 'Отказано в доступе.')
                return
            deadline_str = "Не ограничено."
            if cur_task.deadline != float(inf):
                deadline_str = datetime.datetime.utcfromtimestamp(cur_task.deadline + 3*60*60).strftime("%d-%m-%Y %H:%M:%S")
            bot.send_message(self.id, f'ID: {cur_task.id}\nУсловие: {cur_task.statement}'
                                      f'\nГруппа: {"Публичная" if cur_task.group == "None" else cur_task.group}'
                                      f'\nВидимость посылки: {"Открытая" if cur_task.visible else "Скрытая"}'
                                      f'\nЗасчитываемый результат: {"Лучший" if cur_task.best_or_last == 0 else "Последний"}'
                                      f'\nСрок завершения: {deadline_str}')


activity = LinkedList.LinkedList()
"""Last recent used cache list"""
cache = dict()
"""Last recent used cache hash table"""

max_cache_size = 3  # size of cache
max_afk_time = 10 * 60  # seconds
time_between_checks = 5 * 60  # time between two checks in check_cache() in seconds


def update_cache(id: int, connection: sqlite3.Connection):
    """
    Updates cache if you need to add User there
    :param connection: connection to sql database
    :param id: id of user that you need to add in cache
    """
    cur_time = time.time()
    if id in cache.keys():
        user = cache[id].data[1]
        activity.delete(cache[id])
    else:
        user = database.get_user(id, connection)
    if user is None:
        user = User(id)
        database.add_user(user, connection)
    activity.push((cur_time, user))
    cache[id] = activity.head
    if activity.size > max_cache_size:
        cur_id = activity.tail.data[1].id
        database.update_user_info(activity.tail.data[1], connection)
        cache.pop(cur_id)
        activity.pop()


def cycle_check():
    """Decorator that makes infinity cycle for function with pause between iterations"""
    connection = sqlite3.connect("Files/database.sql")
    """database connection"""
    try:
        while True:
            check_cache(connection)
            #  node = activity.head
            #  while node is not None:
            #      print(node.data[0], node.data[1].id)
            #      node = node.next
            #  print('-' * 20)
            time.sleep(time_between_checks)
    except (KeyboardInterrupt, SystemExit):
        connection.close()
        raise KeyboardInterrupt


def check_cache(connection: sqlite3.Connection):
    """
    Removes elements from cache if they have not been used for a long time
    """
    cur_time = time.time()
    while activity.size and cur_time - activity.tail.data[0] > max_afk_time:
        user = activity.tail.data[1]
        database.update_user_info(user, connection)
        cache.pop(user.id)
        activity.pop()
